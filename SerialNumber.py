import openpyxl
import os
from openpyxl import Workbook


def get_workbook(fileName):
    if not os.path.exists(fileName):
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(fileName)
    return wb


def get_sheet(wb: Workbook, sheetName):
    if not wb.sheetnames.__contains__(sheetName):
        wb.create_sheet(title=sheetName)

    return wb.get_sheet_by_name(sheetName)

def add_2_sheet(content):
    excelFileName = 'test.xlsx'
    sheetName = "testSheetName"

    wb = get_workbook(excelFileName)
    get_sheet(wb, sheetName).append([content])


def get_communication_number():
    communication_type_list = ['通信类型：', 'W - Wifi', 'B - 私有mesh', 'S - Sigmesh', 'Z - Zigbee', 'G - GPRS', 'L - Lora',
                               'N - NB-IoT', 'E - 其他']
    for i in communication_type_list:
        print(i)
    communication_number_list = ['W', 'B', 'S', 'Z', 'G', 'L', 'N', 'E']
    while True:
        communication_type = input("请输入通信类型：\n")
        if not (communication_type in communication_number_list):
            print('输入的类型有误，请重新输入！')
        else:
            break
    return communication_type


def get_voltage_type():
    print("---------------------------------------------------")
    voltage_type_list = ['电压类型：', 'H- 220V', 'L- 110V', 'X- 其他情况']
    voltage_number_list = ['H', 'L', 'X']
    for i in voltage_type_list:
        print(i)
    while True:
        voltage_type = input("请输入电压类型：\n")
        if not (voltage_type in voltage_number_list):
            print('输入的电压类型有误，请重新输入！')
        else:
            break
    return voltage_type


def get_number():
    return '1'


def generate_series_number():
    series_number = get_communication_number() + get_voltage_type() + get_number()
    print('============================================')
    print('生成序列号：' + series_number)
    print('============================================')


while True:
    generate_series_number()

add_2_sheet()