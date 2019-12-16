import openpyxl
import os


class Excel:

    __fileName = 'SerialNumber.xlsx'

    __sheetName = "Numbers"

    def __init__(self):
        self.wb = Excel.__get_workbook(self.__fileName)
        self.ws = Excel.__get_sheet(self.wb, self.__sheetName)

    def get_latest_number(self):
        number = 0
        for row in self.ws.rows:
            for cell in row:
                number = int(cell.value[-6:]) + 1
        return number

    def record(self, serial_number):
        print(serial_number)
        self.ws.append([serial_number])
        self.wb.save(self.__fileName)

    @staticmethod
    def __get_workbook(name):
        if not os.path.exists(name):
            wb = openpyxl.Workbook()
        else:
            wb = openpyxl.load_workbook(name)
        return wb

    @staticmethod
    def __get_sheet(wb, name):
        if not wb.__contains__(name):
            wb.create_sheet(title=name)
        return wb[name]


