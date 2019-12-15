import openpyxl
import os
from openpyxl import Workbook


def get_workbook(fileName):
    if not os.path.exists(fileName):
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(fileName)
    return wb


def get_sheet(wb: Workbook, sheetName):
    if not wb.sheetnames.__contains__(sheetName):
        wb.create_sheet(title=sheetName)

    return wb.get_sheet_by_name(sheetName)


def add_2_sheet(content):
    excelFileName = 'test.xlsx'
    sheetName = "testSheetName"

    wb = get_workbook(excelFileName)
    st = get_sheet(wb, sheetName)
    st.append([content])
    print(st.)

add_2_sheet("hahah")